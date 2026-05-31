#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel数据字典生成模块
导出数据库表结构到Excel文件，支持超链接导航
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from .interface import DatabaseConnection
from .logger import get_logger

logger = get_logger()


class ExcelExporter:
    """Excel导出器"""

    TOC_SHEET_NAME = "目录"
    LINK_COLUMN = 8

    def __init__(self, connection: DatabaseConnection):
        """初始化Excel导出器

        Args:
            connection: 数据库连接对象
        """
        self.connection = connection
        logger.info("Excel导出器初始化完成")

    def _get_db_type(self) -> str:
        """获取数据库类型

        Returns:
            str: 数据库类型 (mysql, postgresql, oracle)
        """
        connection_type = type(self.connection).__name__.lower()
        if 'mysql' in connection_type:
            return 'mysql'
        elif 'postgresql' in connection_type or 'postgres' in connection_type:
            return 'postgresql'
        elif 'oracle' in connection_type:
            return 'oracle'
        return 'mysql'

    def export_selected_tables(self, output_dir: str, selected_tables: List[str],
                               show_progress: bool = True) -> str:
        """导出选中的表到Excel文件

        Args:
            output_dir: 输出目录
            selected_tables: 要导出的表名列表
            show_progress: 是否显示进度

        Returns:
            str: 生成的Excel文件路径
        """
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"data_dictionary_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        logger.info(f"开始导出Excel数据字典: {filename}")
        wb = Workbook()
        toc_sheet = wb.active
        toc_sheet.title = self.TOC_SHEET_NAME

        self._setup_toc_sheet(toc_sheet)

        total = len(selected_tables)
        for idx, table_name in enumerate(selected_tables, start=1):
            try:
                if show_progress:
                    logger.info(f"处理表 {idx}/{total}: {table_name}")

                table_structure = self.connection.get_table_structure(table_name)
                fields = table_structure.get('fields', [])

                safe_sheet_name = self._sanitize_sheet_name(table_name)
                ws = wb.create_sheet(safe_sheet_name)

                self._write_table_header(ws, table_name, toc_sheet, idx)

                for row_idx, field in enumerate(fields, start=2):
                    ws.cell(row=row_idx, column=1, value=field.get('name', ''))
                    ws.cell(row=row_idx, column=2, value=field.get('type', ''))
                    ws.cell(row=row_idx, column=3, value=field.get('length', ''))
                    ws.cell(row=row_idx, column=4, value='否' if field.get('nullable') == False else '是')
                    ws.cell(row=row_idx, column=5, value='是' if field.get('primary_key') else '否')
                    ws.cell(row=row_idx, column=6, value=field.get('default') or '')
                    ws.cell(row=row_idx, column=7, value=field.get('comment') or '')

                self._setup_column_widths(ws)
                self._add_return_link(ws)

                current_date = datetime.now().strftime("%Y-%m-%d")
                toc_sheet.cell(row=idx+1, column=1, value=idx)
                toc_sheet.cell(row=idx+1, column=3, value=table_structure.get('comment', ''))
                toc_sheet.cell(row=idx+1, column=4, value=current_date)
                toc_sheet.cell(row=idx+1, column=5, value="正常")

            except Exception as e:
                error_msg = str(e)
                logger.error(f"导出表 {table_name} 时出错: {error_msg}")
                current_date = datetime.now().strftime("%Y-%m-%d")
                toc_sheet.cell(row=idx+1, column=1, value=idx)
                toc_sheet.cell(row=idx+1, column=2, value=table_name)
                toc_sheet.cell(row=idx+1, column=3, value=f"处理出错: {error_msg}")
                toc_sheet.cell(row=idx+1, column=4, value=current_date)
                toc_sheet.cell(row=idx+1, column=5, value="错误")

        self._setup_toc_sheet_layout(toc_sheet)
        wb.save(filepath)
        logger.info(f"Excel导出完成: {filepath}")
        return filepath

    def _setup_toc_sheet(self, toc_sheet):
        """设置目录页"""
        headers = ["序号", "表名", "注释", "添加时间", "状态"]
        for col_idx, header in enumerate(headers, start=1):
            cell = toc_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def _setup_toc_sheet_layout(self, toc_sheet):
        """设置目录页布局"""
        self._setup_column_widths(toc_sheet)
        toc_sheet.column_dimensions['B'].width = 40

    def _write_table_header(self, ws, table_name: str, toc_sheet, idx: int):
        """写入表头和超链接"""
        ws.cell(row=1, column=1, value=f"表名: {table_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        headers = ["字段名", "数据类型", "长度", "可空", "主键", "默认值", "注释"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        safe_sheet_name = self._sanitize_sheet_name(table_name)
        safe_sheet_name_for_link = safe_sheet_name.replace("'", "''")
        display_text = table_name.replace('"', '""')
        link = f"#'{safe_sheet_name_for_link}'!A1"
        link_formula = f'=HYPERLINK("{link}","{display_text}")'
        toc_sheet.cell(row=idx+1, column=2, value=link_formula)

    def _setup_column_widths(self, ws):
        """设置列宽"""
        widths = [20, 15, 10, 8, 8, 15, 30]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _add_return_link(self, ws):
        """添加返回目录链接"""
        ws.cell(row=1, column=self.LINK_COLUMN, value='=HYPERLINK("#目录!A1","返回目录")')
        ws.cell(row=1, column=self.LINK_COLUMN).font = Font(color="0000FF", underline="single")

    def _sanitize_sheet_name(self, sheet_name: str) -> str:
        """处理特殊表名，替换不允许的字符"""
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
        return sheet_name[:31]

    def _get_table_comment(self, table_name: str) -> str:
        """获取表注释，根据数据库类型使用不同的参数化查询"""
        db_type = self._get_db_type()
        try:
            if db_type == 'mysql':
                query = "SELECT table_comment FROM information_schema.TABLES WHERE table_schema = DATABASE() AND table_name = %s"
                result = self.connection.execute(query, (table_name,))
            elif db_type == 'postgresql':
                schema = getattr(self.connection, 'schema', 'public')
                query = """
                    SELECT obj_description((%s || '.' || %s)::regclass, 'pg_class') AS table_comment
                    FROM information_schema.tables
                    WHERE table_schema = %s AND table_name = %s
                """
                result = self.connection.execute(query, (schema, table_name, schema, table_name))
            elif db_type == 'oracle':
                query = """
                    SELECT comments
                    FROM user_tab_comments
                    WHERE table_name = upper(:1)
                """
                result = self.connection.execute(query, [table_name])
            else:
                query = "SELECT table_comment FROM information_schema.TABLES WHERE table_schema = DATABASE() AND table_name = %s"
                result = self.connection.execute(query, (table_name,))

            if result and len(result) > 0:
                if db_type == 'mysql':
                    return result[0].get('table_comment', '')
                elif db_type == 'postgresql':
                    return result[0].get('table_comment', '') if isinstance(result[0], dict) else (result[0][0] if result[0] else '')
                elif db_type == 'oracle':
                    return result[0][0] if result[0] else ''
            return ""
        except Exception as e:
            logger.debug(f"获取表注释失败 {table_name}: {str(e)}")
            return ""
